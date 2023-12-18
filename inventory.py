#!/usr/bin/env python

import yaml
import configparser
import tomllib
from yaml import SafeDumper
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter

config_file = "settings.ini"

SafeDumper.add_representer(
    type(None),
    lambda dumper, value: dumper.represent_scalar(u'tag:yaml.org,2002:null', '')
  )

def main():
    config = load_config(config_file)
    wb = load_workbook(config['file']) 
    sheet = wb[config['sheet']]

    hostname_col = column_index_from_string(config['hostname_col'])-1
    
    group_by_cols = [1,2]
    variable_cols = [5,6,7]
    inventory = sheet_to_inventory(group_by_cols,hostname_col,variable_cols,sheet)

    print(yaml.safe_dump(inventory))

def sheet_to_inventory(group_by_cols,hostname_col,variable_cols,sheet):
    rows = list(sheet.rows)
    groups = {}

    for row in rows[1:]:
        group = row[group_by_cols[0]].value + "_" + row[group_by_cols[1]].value
        host = row[hostname_col].value
        if group not in groups:
            groups[group] = {"hosts": {}}
        groups[group]["hosts"][host] = {}
        for col in variable_cols:
            #print(host + " " + sheet[get_column_letter(col+1) + "1"].value, row[col].value)
            if row[col].value is not None:
                groups[group]["hosts"][host][sheet[get_column_letter(col+1) + "1"].value] = row[col].value

    return groups

def load_config(config_file):
    with open(config_file, "rb") as f:
        data = tomllib.load(f)
    return data

if __name__ == "__main__":
    main()
